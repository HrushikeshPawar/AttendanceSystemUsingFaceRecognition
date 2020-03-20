from openpyxl import Workbook, load_workbook
import time
import os
import sqlite3
import datetime

names = []
rolls = []
for text in os.listdir("Datasets/"):
    NR = text.split("_")
    names.append(NR[0])
    rolls.append(NR[1])

#get current date
currentDate = time.strftime("%d/%m/%y")
day = time.strftime('%a')
#create a workbook and add a worksheet
if(os.path.exists("Attendance.xlsx")):
    wb = load_workbook(filename = "Attendance.xlsx")
else:
    wb = Workbook()

dest_filename = 'Attendance.xlsx'
#creating worksheet and giving names to column
ws1 = wb.active

ws1.title = "Real Analysis"
ws1.append(('Roll Number', 'Name', ''))
ws1.append(('', '', ''))


ws = wb.get_sheet_by_name('Real Analysis')  # Work Sheet
column = ws['A']  # Column
column_list = [column[x].value for x in range(len(column))]


#entering students information from database
for i in range(len(names)):
    if rolls[i] in column_list : continue
    else: ws1.append((rolls[i], names[i]))


#saving the file
wb.save(filename = dest_filename)
