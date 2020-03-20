from openpyxl import Workbook, load_workbook
import time
from PyQt5 import QtCore, QtGui, QtWidgets
import os


def Register_Subject(sub):
    wb = load_workbook(filename = "Attendance.xlsx")
    sheet = wb.sheetnames
    if sub in sheet:
        return 0
    else:
        ws = wb.create_sheet("Sheet_A")
        ws.title = sub

        names = []
        rolls = []
        for text in os.listdir("Datasets/"):
            NR = text.split("_")
            names.append(NR[0])
            rolls.append(NR[1])


        ws.append(('Roll Number', 'Name', ''))
        ws.append(('', '', ''))

        for i in range(len(names)):
            ws.append((rolls[i], names[i]))

        wb.save(filename = "Attendance.xlsx")
        return 1


def Register_Student():
    wb = load_workbook(filename = "Attendance.xlsx")

    names = []
    rolls = []
    for text in os.listdir("Datasets/"):
        NR = text.split("_")
        names.append(NR[0])
        rolls.append(NR[1])

    for ws in wb:
        column = ws['A']  
        column_list = [column[x].value for x in range(len(column))]

        for i in range(len(names)):
            if rolls[i] in column_list : continue
            else: ws.append((rolls[i], names[i]))

    wb.save(filename = "Attendance.xlsx")
