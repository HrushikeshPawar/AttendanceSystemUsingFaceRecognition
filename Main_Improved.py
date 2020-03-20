#Import all required libraries and scripts
from PyQt5 import QtCore, QtGui, QtWidgets
from openpyxl import Workbook, load_workbook
import time
import math
import os
import os.path
import pickle
from PIL import Image, ImageDraw
import face_recognition
from face_recognition.face_recognition_cli import image_files_in_folder
from sklearn import neighbors

from Register_Excel import Register_Student, Register_Subject
from Dataset_Improved import Data
from Face_Recognition_knn import predict, show_prediction_labels_on_image
from Train_knn import train





#class - Hold everything together, giving it a well defined structure
class Ui_MainWindow(object):

    #This create the GUI - Code auto generated from the design software
    def setupUi(self, MainWindow):
        MainWindow.setObjectName("MainWindow")
        MainWindow.resize(519, 432)
        self.centralwidget = QtWidgets.QWidget(MainWindow)
        self.centralwidget.setObjectName("centralwidget")
        self.label = QtWidgets.QLabel(self.centralwidget)
        self.label.setGeometry(QtCore.QRect(10, 10, 481, 51))
        font = QtGui.QFont()
        font.setFamily("MathJax_SansSerif")
        font.setPointSize(20)
        font.setBold(True)
        font.setWeight(75)
        self.label.setFont(font)
        self.label.setObjectName("label")
        self.tabWidget = QtWidgets.QTabWidget(self.centralwidget)
        self.tabWidget.setGeometry(QtCore.QRect(26, 69, 461, 291))
        font = QtGui.QFont()
        font.setPointSize(15)
        self.tabWidget.setFont(font)
        self.tabWidget.setTabShape(QtWidgets.QTabWidget.Rounded)
        self.tabWidget.setObjectName("tabWidget")
        self.tab2 = QtWidgets.QWidget()
        self.tab2.setObjectName("tab2")
        self.comboBox = QtWidgets.QComboBox(self.tab2)
        self.comboBox.setGeometry(QtCore.QRect(220, 5, 151, 30))
        self.comboBox.setEditable(False)
        self.comboBox.setCurrentText("")
        self.comboBox.setObjectName("comboBox")
        self.label_5 = QtWidgets.QLabel(self.tab2)
        self.label_5.setGeometry(QtCore.QRect(30, 10, 111, 20))
        font = QtGui.QFont()
        font.setPointSize(12)
        self.label_5.setFont(font)
        self.label_5.setObjectName("label_5")
        self.loadImg = QtWidgets.QPushButton(self.tab2)
        self.loadImg.setGeometry(QtCore.QRect(20, 100, 151, 31))
        self.loadImg.setObjectName("loadImg")
        self.listWidget = QtWidgets.QListWidget(self.tab2)
        self.listWidget.setGeometry(QtCore.QRect(180, 70, 256, 101))
        self.listWidget.setObjectName("listWidget")
        self.markAttendance = QtWidgets.QPushButton(self.tab2)
        self.markAttendance.setGeometry(QtCore.QRect(100, 190, 251, 41))
        self.markAttendance.setObjectName("markAttendance")
        self.tabWidget.addTab(self.tab2, "")
        self.tab1 = QtWidgets.QWidget()
        font = QtGui.QFont()
        font.setPointSize(11)
        self.tab1.setFont(font)
        self.tab1.setObjectName("tab1")
        self.toolBox = QtWidgets.QToolBox(self.tab1)
        self.toolBox.setGeometry(QtCore.QRect(20, 20, 411, 231))
        font = QtGui.QFont()
        font.setPointSize(14)
        self.toolBox.setFont(font)
        self.toolBox.setObjectName("toolBox")
        self.page1 = QtWidgets.QWidget()
        self.page1.setGeometry(QtCore.QRect(0, 0, 411, 159))
        self.page1.setObjectName("page1")
        self.SubReg = QtWidgets.QPushButton(self.page1)
        self.SubReg.setGeometry(QtCore.QRect(210, 60, 151, 31))
        self.SubReg.setObjectName("SubReg")
        self.SubName = QtWidgets.QLineEdit(self.page1)
        self.SubName.setGeometry(QtCore.QRect(0, 60, 181, 31))
        self.SubName.setObjectName("SubName")
        self.label_2 = QtWidgets.QLabel(self.page1)
        self.label_2.setGeometry(QtCore.QRect(30, 30, 131, 20))
        font = QtGui.QFont()
        font.setPointSize(12)
        self.label_2.setFont(font)
        self.label_2.setObjectName("label_2")
        self.toolBox.addItem(self.page1, "")
        self.page2 = QtWidgets.QWidget()
        self.page2.setGeometry(QtCore.QRect(0, 0, 411, 159))
        self.page2.setObjectName("page2")
        self.label_3 = QtWidgets.QLabel(self.page2)
        self.label_3.setGeometry(QtCore.QRect(30, 20, 71, 20))
        font = QtGui.QFont()
        font.setPointSize(12)
        self.label_3.setFont(font)
        self.label_3.setObjectName("label_3")
        self.label_4 = QtWidgets.QLabel(self.page2)
        self.label_4.setGeometry(QtCore.QRect(30, 60, 71, 20))
        font = QtGui.QFont()
        font.setPointSize(12)
        self.label_4.setFont(font)
        self.label_4.setObjectName("label_4")
        self.name = QtWidgets.QLineEdit(self.page2)
        self.name.setGeometry(QtCore.QRect(150, 10, 211, 31))
        self.name.setAlignment(QtCore.Qt.AlignCenter)
        self.name.setObjectName("name")
        self.roll = QtWidgets.QLineEdit(self.page2)
        self.roll.setGeometry(QtCore.QRect(150, 60, 211, 31))
        self.roll.setAlignment(QtCore.Qt.AlignCenter)
        self.roll.setObjectName("roll")
        self.StudentReg = QtWidgets.QPushButton(self.page2)
        self.StudentReg.setGeometry(QtCore.QRect(250, 100, 151, 31))
        self.StudentReg.setObjectName("StudentReg")
        self.Train = QtWidgets.QPushButton(self.page2)
        self.Train.setGeometry(QtCore.QRect(20, 100, 151, 31))
        self.Train.setObjectName("Train")
        self.toolBox.addItem(self.page2, "")
        self.tabWidget.addTab(self.tab1, "")
        MainWindow.setCentralWidget(self.centralwidget)
        self.menubar = QtWidgets.QMenuBar(MainWindow)
        self.menubar.setGeometry(QtCore.QRect(0, 0, 519, 22))
        self.menubar.setObjectName("menubar")
        MainWindow.setMenuBar(self.menubar)
        self.statusbar = QtWidgets.QStatusBar(MainWindow)
        self.statusbar.setObjectName("statusbar")
        MainWindow.setStatusBar(self.statusbar)

        self.retranslateUi(MainWindow)
        self.tabWidget.setCurrentIndex(0)
        self.comboBox.setCurrentIndex(-1)
        self.toolBox.setCurrentIndex(1)
        QtCore.QMetaObject.connectSlotsByName(MainWindow)
        # ------ The auto - generated code ends ------


        wb = load_workbook(filename = "Attendance.xlsx") #Load the excel file for use
        sheet = wb.sheetnames #Get the list of all the sheets in the file (Sub list)
        self.comboBox.addItem("---Subjects---") #Enter this list in th drop down box
        self.comboBox.addItems(sheet)

        #All the click events
        self.SubReg.clicked.connect(self.RegSub)  #Click the button to Register Subject
        self.StudentReg.clicked.connect(self.RegStu) #Click the button to Register Students
        self.loadImg.clicked.connect(self.Insert)  #Click the button to Insert Images
        self.markAttendance.clicked.connect(self.Attendance) #Click the button to Mark Attendance
        self.Train.clicked.connect(self.Training) #CLick the button to Train the knn model

    def retranslateUi(self, MainWindow):  #Part of Auto - Generated code
        _translate = QtCore.QCoreApplication.translate
        MainWindow.setWindowTitle(_translate("MainWindow", "Attendance"))
        self.label.setText(_translate("MainWindow", "   Face Recognition Attendance System"))
        self.label_5.setText(_translate("MainWindow", "Select Subject : "))
        self.loadImg.setText(_translate("MainWindow", "Insert"))
        self.markAttendance.setText(_translate("MainWindow", "Mark Attendance"))
        self.tabWidget.setTabText(self.tabWidget.indexOf(self.tab2), _translate("MainWindow", "Attendance"))
        self.SubReg.setText(_translate("MainWindow", "Register"))
        self.label_2.setText(_translate("MainWindow", "Subject Name"))
        self.toolBox.setItemText(self.toolBox.indexOf(self.page1), _translate("MainWindow", "Subject Registration"))
        self.label_3.setText(_translate("MainWindow", "  Name  : "))
        self.label_4.setText(_translate("MainWindow", "Roll No. : "))
        self.StudentReg.setText(_translate("MainWindow", "Register"))
        self.Train.setText(_translate("MainWindow", "Train"))
        self.toolBox.setItemText(self.toolBox.indexOf(self.page2), _translate("MainWindow", "Student Registration"))
        self.tabWidget.setTabText(self.tabWidget.indexOf(self.tab1), _translate("MainWindow", "Registration"))

    def RegSub(self): #User-Defined-Fuction (UFD) to register subject
        sub = self.SubName.text() #Get the name of the subject entered in th textbox
        sub = sub.strip() #Remove all the spaces before or afte the word, if any
        if(sub == ""): #If the input is empty or just space give the error
            #A code to generate a message box to show Error.
            # This same code is used everywhere to show message box whenever required
            self.msg = QtWidgets.QMessageBox()
            self.msg.setIcon(QtWidgets.QMessageBox.Critical)
            self.msg.setText("Invalid Subject Name")
            self.msg.setWindowTitle("Error")
            self.msg.setStandardButtons(QtWidgets.QMessageBox.Ok)
            self.msg.show()
        else:
            N = Register_Subject(sub) #Checks if the subject is already registered
            if(N==0): #If yes, give an error using a msgbox as above and stop
                self.msg = QtWidgets.QMessageBox()
                self.msg.setIcon(QtWidgets.QMessageBox.Warning)
                self.msg.setText("Subject Already Registered !")
                self.msg.setWindowTitle("Error")
                self.msg.setStandardButtons(QtWidgets.QMessageBox.Ok)
                self.msg.show()
            else: #if NO, then give completion msg
                self.msg = QtWidgets.QMessageBox()
                self.msg.setIcon(QtWidgets.QMessageBox.Information)
                self.msg.setText("Subject Registered Sucessfully !")
                self.msg.setWindowTitle("Sucess")
                self.msg.setStandardButtons(QtWidgets.QMessageBox.Ok)
                self.msg.show()
                self.SubName.setText("") #empty the msgbox
                self.comboBox.addItem(sub) #add subject to the excelfile as new sheet

    def RegStu(self): #UFD for student registeration
        name = self.name.text() #Take text from textbox in front of name
        name = name.strip() #Remove spaces if any
        roll = self.roll.text() #Take roll no. as text
        roll = roll.strip() #Removes spaces from that as well
        if(name == "" or roll == ""): #Check if any of the input is empty or just spaces if yes, give error and stop
            self.msg = QtWidgets.QMessageBox()
            self.msg.setIcon(QtWidgets.QMessageBox.Critical)
            self.msg.setText("Invalid Student Name or Roll No.")
            self.msg.setWindowTitle("Error")
            self.msg.setStandardButtons(QtWidgets.QMessageBox.Ok)
            self.msg.show()
            return

        for file in os.listdir("Datasets/"): #Create a list of all the names and roll in the Dataset
            text = file.split("_")
            Roll = text[1]
            if(roll == Roll): #Check if the roll number is already registered, if yes give error and stop
                self.msg = QtWidgets.QMessageBox()
                self.msg.setIcon(QtWidgets.QMessageBox.Critical)
                self.msg.setText("Roll No. already exists !")
                self.msg.setWindowTitle("Error")
                self.msg.setStandardButtons(QtWidgets.QMessageBox.Ok)
                self.msg.show()
                return

        #Path to save the Data
        path = "Datasets/"+ name + "_" + roll

        #Check if the username already exists. If not, store the images here.
        if os.path.exists(path) == True: #If Exists the print error and stop
            self.msg = QtWidgets.QMessageBox()
            self.msg.setIcon(QtWidgets.QMessageBox.Warning)
            self.msg.setText("This Student Name and Roll No. is Already Registered !")
            self.msg.setWindowTitle("Error")
            self.msg.setStandardButtons(QtWidgets.QMessageBox.Ok)
            self.msg.show()
        else: #If NO, then create folder with name and roll to save the image files in Dataset
            os.mkdir(path)
            self.msg = QtWidgets.QMessageBox() #Create a msgbox to display the Instructions
            self.msg.setIcon(QtWidgets.QMessageBox.Information)
            self.msg.setText("Please read the intructions given below.")
            self.msg.setWindowTitle("Student Registration")
            self.msg.setStandardButtons(QtWidgets.QMessageBox.Ok | QtWidgets.QMessageBox.Cancel)
            self.msg.setDetailedText(" [Instructions] \n 1. Wait for camera to start.\n 2. Press 'k' to click a photo. \n 3. Click 10 images for each student. \n 4. [IMP] Click on 'Train' after all students are Registered.")
            ret = self.msg.exec()
            if(ret == QtWidgets.QMessageBox.Ok): #If the OK button is pressed then only procced further
                Data(name,roll,path) #Call the Data function for Dataset.py file which opens the camera and creates the dataset
                Register_Student() #Add name and roll to the excelfile
                self.msg = QtWidgets.QMessageBox() #Show a msg box after the work is done
                self.msg.setIcon(QtWidgets.QMessageBox.Information)
                self.msg.setText("Student Registered Successfully !")
                self.msg.setWindowTitle("Student Registration")
                self.msg.setStandardButtons(QtWidgets.QMessageBox.Ok)
                self.msg.show()
                self.name.setText("") #Remove texts from the textboxs
                self.roll.setText("")

    def Insert(self): #UDF to insert the img file in the mail window
        file, _ = QtWidgets.QFileDialog.getOpenFileName(None,caption="Select Image",
        directory="/home/hrushikesh/Desktop/Face_Recognition_Attendance_System",
        filter="Image Files (*.png *.jpg *.jpeg *.bmp)")
        self.listWidget.addItem(file) #Add the file name to the listbox

    def Attendance(self): #UDF to mark the attendance
        img = []
        sub = self.comboBox.currentText() #Get the selected subject from dropdown list
        if(self.listWidget.count()==0): #If no image is selected then give an error
            self.msg = QtWidgets.QMessageBox()
            self.msg.setIcon(QtWidgets.QMessageBox.Warning)
            self.msg.setText("Image not selected !")
            self.msg.setWindowTitle("Attendance")
            self.msg.setStandardButtons(QtWidgets.QMessageBox.Ok)
            self.msg.show()
        elif(sub == "---Subjects---"): #If no image is selected give an error
            self.msg = QtWidgets.QMessageBox()
            self.msg.setIcon(QtWidgets.QMessageBox.Warning)
            self.msg.setText("Subject not selected !")
            self.msg.setWindowTitle("Attendance")
            self.msg.setStandardButtons(QtWidgets.QMessageBox.Ok)
            self.msg.show()
        else:
            for i in range(self.listWidget.count()):
                img.append(self.listWidget.item(i).text()) #Get the list of the imag file

            for image_file in img: #Select one image at a time from the list of images
                file = image_file.split("/")
                file_name = file[len(file)-1]
                print("\n [INFO] Looking for faces in {}".format(file_name))

                # Find all people in the image using a trained classifier model
                # Note: You can pass in either a classifier file name or a classifier model instance
                predictions = predict(image_file, model_path="Output_Files/trained_knn_model.clf")

                # Get Present Day and Date
                date = time.strftime("%d/%m/%y")
                day = time.strftime('%a')

                # Open The subject attendance sheet from excelfile
                wb = load_workbook(filename = "Attendance.xlsx")
                ws = wb[sub]
                i = 3
                while(True): #Print the day and date in the next empty column
                    i += 1
                    cell = ws.cell(row=1,column=i)
                    cell = cell.value
                    if cell == None:
                        ws.cell(row=1,column=i).value = date
                        ws.cell(row=2,column=i).value = day
                        break

                # Get the place of name and roll number in the excel sheet
                column = ws['A']
                column_list = [column[x].value for x in range(len(column))]
                row = ws[1]
                row_list = [row[x].value for x in range(len(row))]

                for text, (top, right, bottom, left) in predictions:
                    if text != "unknown": #If the person is recognized, do as follow
                        NR = text.split("_")
                        name = NR[0]
                        roll = NR[1]
                        print("\t - Found {}".format(name, left, top))
                        x = column_list.index(roll)
                        y = row_list.index(date)
                        c = ws.cell(row=x+1,column=y+1)
                        c.value = "Present" # print present before his name in the right column
                    else:
                        name = text #if the face is not recognized, skip
                        print("\t - Found {}".format(name))

                # SAVE the excelfile once everthing is done
                wb.save(filename = "Attendance.xlsx")

                #Print a msgbox saying the attendance has been marked
                self.msg = QtWidgets.QMessageBox()
                self.msg.setIcon(QtWidgets.QMessageBox.Information)
                text = "Attendance marked Sucessfully for {}".format(sub)
                self.msg.setText(text)
                self.msg.setWindowTitle("Attendance")
                self.msg.setStandardButtons(QtWidgets.QMessageBox.Ok)
                self.msg.show()
                self.comboBox.setCurrentIndex(0)
                self.listWidget.clear() #Clear the list of images

    def Training(self): #UDF for training the dataset
        train() #Call a function from the train file


if __name__ == "__main__":  #Starts the main program and creates the window
    import sys
    app = QtWidgets.QApplication(sys.argv)
    MainWindow = QtWidgets.QMainWindow()
    ui = Ui_MainWindow()
    ui.setupUi(MainWindow)
    MainWindow.show()
    sys.exit(app.exec_())
